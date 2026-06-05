import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def export_to_excel(resources_by_type, resource_group, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = resource_group
    # Define columns
    columns = [
        "Nome Completo Risorsa", "Hostname", "IP", "Port", "DB Name", "Admin Username", "SSH key", "Modalità", "Environment", "Application ID", "Application Name", "Application Group", "Resource Function", "Resource Role", "Spazio", "Tipologia Macchina", "Subnet", "Network Security Group", "Schedule", "Backup Strategy", "OS", "Note"
    ]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def get_tag(tags, key):
        if not tags:
            return None
        return tags.get(key) or tags.get(key.lower()) or tags.get(key.upper())

    # Helper to add a row
    def add_row(resource, hostname=None, ip=None):
        tags = resource.get('tags', {})
        resource_type = resource.get('type')
        tipologia_macchina = None
        spazio = ''
        admin_username = resource.get('adminLogin')
        excel_hostname = hostname
        excel_ip = ip
        excel_subnet = ''
        note = ''
        # Explicit mappings by resource type
        if resource_type == 'Microsoft.CognitiveServices/accounts':
            tipologia_macchina = resource.get('pricing_tier')
        elif resource_type == 'Microsoft.DBforPostgreSQL/flexibleServers':
            tipologia_macchina = resource.get('sku', {}).get('name')
            storage_gb = resource.get('storage', {}).get('storageSizeGB')
            spazio = f"{storage_gb} GB" if storage_gb else ''
            admin_username = resource.get('adminLogin')
            excel_hostname = resource.get('fullyQualifiedDomainName')
            # Compose subnet column as VN/SUBNET (IP range)
            subnet_id = None
            vnet_name = None
            subnet_name = None
            ip_range = None
            # Try to parse subnet info from delegatedSubnetResourceId if present
            if resource.get('subnet') and isinstance(resource.get('subnet'), str):
                # If subnet is just the IP range, we need to parse the VN/SUBNET from delegatedSubnetResourceId
                subnet_id = resource.get('delegatedSubnetResourceId') or ''
                if subnet_id:
                    parts = subnet_id.split('/')
                    if 'virtualNetworks' in parts and 'subnets' in parts:
                        vnet_name = parts[parts.index('virtualNetworks') + 1]
                        subnet_name = parts[parts.index('subnets') + 1]
                ip_range = resource.get('subnet')
                if vnet_name and subnet_name and ip_range:
                    excel_subnet = f"{vnet_name}/{subnet_name} ({ip_range})"
                elif ip_range:
                    excel_subnet = ip_range
            elif resource.get('subnet') and isinstance(resource.get('subnet'), dict):
                # If subnet is a dict, try to extract info
                vnet_name = resource.get('subnet').get('vnet')
                subnet_name = resource.get('subnet').get('subnet')
                ip_range = resource.get('subnet').get('addressPrefix')
                if vnet_name and subnet_name and ip_range:
                    excel_subnet = f"{vnet_name}/{subnet_name} ({ip_range})"
                elif ip_range:
                    excel_subnet = ip_range
        elif resource_type == 'Microsoft.Compute/virtualMachines':
            tipologia_macchina = resource.get('vmSize')
            excel_ip = resource.get('privateIpAddress')
        elif resource_type == 'Microsoft.Compute/disks':
            tipologia_macchina = resource.get('sku')
            disk_size = resource.get('diskSizeGB')
            spazio = f"{disk_size} GB" if disk_size else ''
            # Set Note column for disks with LUN
            note = ''
            if resource.get('lun') is not None:
                note = f"LUN: {resource.get('lun')}"
        elif resource_type == 'Microsoft.Network/networkInterfaces':
            excel_ip = resource.get('privateIpAddresses')
        else:
            tipologia_macchina = resource.get('vmSize')
        # If ip is a list, join as comma-separated string
        if isinstance(excel_ip, list):
            ip_str = ', '.join(excel_ip)
        else:
            ip_str = excel_ip
        # Set Note column for disks with LUN (moved to disk elif above)
        ws.append([
            resource.get('name'),
            excel_hostname,
            ip_str,
            '',  # Port
            '',  # DB Name
            admin_username,
            '',  # SSH key
            '',  # Modalità
            get_tag(tags, 'Environment'),
            get_tag(tags, 'Application_ID'),
            get_tag(tags, 'Application_Name'),
            get_tag(tags, 'Application_Group'),
            get_tag(tags, 'Resource_Function'),
            get_tag(tags, 'Resource_Role'),
            spazio,
            tipologia_macchina,
            excel_subnet if resource_type == 'Microsoft.DBforPostgreSQL/flexibleServers' else '',
            '',  # Network Security Group
            get_tag(tags, 'Schedule'),
            get_tag(tags, 'Backup_Strategy'),
            resource.get('osType'),
            note
        ])

    for rtype, resources in resources_by_type.items():
        for resource in resources:
            pes = resource.get('privateEndpoints', [])
            if pes:
                for pe in pes:
                    ip = pe.get('ip_addresses') if pe.get('ip_addresses') else None
                    add_row(resource, hostname=pe.get('hostname'), ip=ip)
            else:
                add_row(resource)
    wb.save(output_path)
